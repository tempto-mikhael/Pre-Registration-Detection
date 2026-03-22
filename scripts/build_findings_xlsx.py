#!/usr/bin/env python3
"""
Build a consolidated findings XLSX from the original XLSX + pipeline outputs.

New / changed columns vs. the original journal_articles_with_pap_2025-03-14.xlsx:
  - no_data        : auto-computed where not already set by an RA (1 = theoretical/
                     review, i.e. all type_* cols are 0/null); existing human labels
                     (0 or 1) are preserved unchanged.
  - auto_prereg    : pipeline keyword-scan result (0/1) from pdf_scan_results.csv
  - auto_link_prereg : best preregistration link found in the PDF (string or empty)
  - ai_prereg      : LLM verdict  True / False / None (no verdict run)
  - ai_confidence  : numeric 0–1  high→0.9 | medium→0.6 | low→0.25 | missing→None
  - pipeline_prereg: COMBINED pipeline decision (1 = preregistered according to our
                     data): 1 if auto_link_prereg is non-empty OR ai_prereg is True;
                     0 if ai_prereg is False AND no link found;
                     None if no verdict and no link evidence.

Output: output/findings_<YYYY-MM-DD>.xlsx
"""

import csv
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not found — run: pip install openpyxl")

ROOT         = Path(__file__).parent.parent
OUTPUT_DIR   = ROOT / "output"
SOURCE_XLSX  = ROOT / "journal_articles_with_pap_2025-03-14.xlsx"
VERDICTS_CSV = OUTPUT_DIR / "llm_gemini_verdicts.csv"
SCAN_CSV     = OUTPUT_DIR / "pdf_scan_results.csv"
LINKS_CSV    = OUTPUT_DIR / "pdf_scan_prereg_links_dedup.csv"
OUT_XLSX     = OUTPUT_DIR / f"findings_pipeline_{date.today()}.xlsx"

CONFIDENCE_MAP = {
    "high":   0.9,
    "medium": 0.6,
    "low":    0.25,
}

TYPE_COLS = ["type_lab", "type_field", "type_online",
             "type_survey", "type_obs", "type_repl", "type_other"]

# ── Header / style constants ──────────────────────────────────────────────────
NEW_COL_FILL      = PatternFill("solid", fgColor="FFF2CC")   # light yellow
AI_COL_FILL       = PatternFill("solid", fgColor="DDEEFF")   # light blue
PIPELINE_COL_FILL = PatternFill("solid", fgColor="D9EAD3")   # light green
HEADER_FONT       = Font(bold=True)
WRAP              = Alignment(wrap_text=True)


def load_csv_by_filename(path: Path, fname_col: str = "filename") -> dict:
    """Return {filename: row_dict} from any CSV that has a filename column."""
    if not path.exists():
        print(f"WARNING: file not found at {path}")
        return {}
    out = {}
    with open(path, encoding="utf-8", errors="ignore") as f:
        for row in csv.DictReader(f):
            fname = (row.get(fname_col) or "").strip()
            if fname:
                out[fname] = row
    return out


def load_scan(path: Path) -> dict:
    """Return {filename: row_dict} from pdf_scan_results.csv."""
    return load_csv_by_filename(path, fname_col="filename")


def load_best_links(path: Path) -> dict:
    """Return {filename: best_link_url} from the dedup links CSV.
    Uses the first (highest-quality) link per paper.
    """
    if not path.exists():
        print(f"WARNING: links file not found at {path}")
        return {}
    out = {}
    with open(path, encoding="utf-8", errors="ignore") as f:
        for row in csv.DictReader(f):
            fname = (row.get("filename") or "").strip()
            if fname and fname not in out:
                # prefer 'url' col, fall back to 'link'
                url = (row.get("url") or row.get("link") or "").strip()
                if url:
                    out[fname] = url
    return out


def is_type_empty(row_values: tuple, type_idxs: list) -> bool:
    """Return True if all type_* columns are 0 or null."""
    for i in type_idxs:
        v = row_values[i]
        if v not in (None, 0, "", "0"):
            return False
    return True


def to_bool_or_none(val: str):
    if val is None:
        return None
    v = val.strip().lower()
    if v == "true":
        return True
    if v == "false":
        return False
    return None


def main():
    print(f"Reading source: {SOURCE_XLSX}")
    src_wb = openpyxl.load_workbook(str(SOURCE_XLSX), read_only=True, data_only=True)
    src_ws = src_wb.active

    all_rows = list(src_ws.iter_rows(values_only=True))
    src_wb.close()

    keyword_label_row = list(all_rows[0])
    header_row        = list(all_rows[1])
    data_rows         = all_rows[2:]
    print(f"Source: {len(data_rows)} data rows, {len(header_row)} columns")

    # Column index lookups (original)
    no_data_idx = header_row.index("no_data")
    pdf_idx     = header_row.index("pdf")
    type_idxs   = [header_row.index(c) for c in TYPE_COLS]

    # Load pipeline data sources
    verdicts   = load_csv_by_filename(VERDICTS_CSV)
    scan       = load_scan(SCAN_CSV)
    link_rows  = load_csv_by_filename(LINKS_CSV)   # dedup enriched links
    print(f"Loaded: {len(verdicts)} LLM verdicts, {len(scan)} scan rows, "
          f"{len(link_rows)} enriched link rows")

    # ── Build output workbook ─────────────────────────────────────────────────
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "findings"

    # New columns appended after the original 53
    NEW_COLS = [
        "auto_prereg",       # pipeline keyword scan result
        "auto_link_prereg",  # direct link found in PDF text
        "enriched_links",    # all links found via enrichment (may be multi)
        "ai_prereg",         # LLM verdict True/False/None
        "ai_confidence",     # numeric 0–1
        "pipeline_prereg",   # COMBINED: link OR LLM → 1/0/None
    ]

    # ── Row 1: keyword label row (pad with blanks) ────────────────────────────
    out_ws.append(keyword_label_row + [None] * len(NEW_COLS))

    # ── Row 2: headers ────────────────────────────────────────────────────────
    header_out = header_row + NEW_COLS
    out_ws.append(header_out)

    # Style header row
    for col_idx, col_name in enumerate(header_out, start=1):
        cell = out_ws.cell(row=2, column=col_idx)
        cell.font = HEADER_FONT
        if col_name == "no_data":
            cell.fill = NEW_COL_FILL
        elif col_name in ("auto_prereg", "auto_link_prereg", "enriched_links"):
            cell.fill = NEW_COL_FILL
        elif col_name in ("ai_prereg", "ai_confidence"):
            cell.fill = AI_COL_FILL
        elif col_name == "pipeline_prereg":
            cell.fill = PIPELINE_COL_FILL

    # ── Data rows ─────────────────────────────────────────────────────────────
    nd_computed   = 0
    verdict_match = 0
    scan_match    = 0
    link_match    = 0

    for raw_row in data_rows:
        row = list(raw_row)

        # no_data: preserve human label; auto-compute only if None
        current_nd = row[no_data_idx]
        if current_nd not in (0, 1):
            row[no_data_idx] = 1 if is_type_empty(row, type_idxs) else 0
            nd_computed += 1

        pdf_val = (row[pdf_idx] or "").strip()

        # ── Scan columns ──────────────────────────────────────────────────────
        s = scan.get(pdf_val)
        if s:
            scan_match += 1
            auto_prereg_val     = int(s.get("auto_prereg", 0) or 0)
            auto_link_prereg_val = (s.get("auto_link_prereg") or "").strip()
        else:
            auto_prereg_val      = None
            auto_link_prereg_val = ""

        # ── Enriched links ────────────────────────────────────────────────────
        lr = link_rows.get(pdf_val)
        if lr:
            link_match += 1
            enriched = (lr.get("all_found_links") or "").strip()
            # if scan didn't give a direct link, try enrichment best link
            if not auto_link_prereg_val:
                auto_link_prereg_val = (lr.get("auto_link_prereg") or "").strip()
        else:
            enriched = ""

        # ── LLM verdict ───────────────────────────────────────────────────────
        verdict = verdicts.get(pdf_val)
        if verdict:
            verdict_match += 1
            ai_prereg     = to_bool_or_none(verdict.get("llm_prereg", ""))
            conf_str      = (verdict.get("llm_confidence") or "").strip().lower()
            ai_confidence = CONFIDENCE_MAP.get(conf_str)
        else:
            ai_prereg     = None
            ai_confidence = None

        # ── Pipeline combined decision ─────────────────────────────────────────
        has_link = bool(auto_link_prereg_val or enriched)
        if has_link or ai_prereg is True:
            pipeline_prereg = 1
        elif ai_prereg is False and not has_link:
            pipeline_prereg = 0
        else:
            pipeline_prereg = None   # no evidence either way

        row_out = row + [
            auto_prereg_val,
            auto_link_prereg_val or None,
            enriched or None,
            ai_prereg,
            ai_confidence,
            pipeline_prereg,
        ]
        out_ws.append(row_out)

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "ra": 8, "flag": 6, "comment": 30, "no_data": 9,
        "id": 6, "file_name": 50, "file_title": 50, "journal": 35,
        "pdf": 40, "prereg": 8, "link_prereg": 40,
        "use_aearct": 10, "use_osf": 8, "use_aspredicted": 14, "use_other": 10,
        "aearct_pap": 10,
        "type_lab": 9, "type_field": 10, "type_online": 10,
        "type_survey": 10, "type_obs": 9, "type_repl": 9, "type_other": 10,
        "auto_prereg": 12, "auto_link_prereg": 45, "enriched_links": 60,
        "ai_prereg": 12, "ai_confidence": 14, "pipeline_prereg": 16,
    }
    for col_idx, col_name in enumerate(header_out, start=1):
        width = col_widths.get(col_name, 7)
        out_ws.column_dimensions[get_column_letter(col_idx)].width = width

    out_ws.freeze_panes = "A3"

    # ── Save ─────────────────────────────────────────────────────────────────
    out_wb.save(str(OUT_XLSX))

    print(f"\nDone.")
    print(f"  no_data auto-computed for:        {nd_computed} rows")
    print(f"  Scan data merged:                 {scan_match} rows")
    print(f"  Enriched link data merged:        {link_match} rows")
    print(f"  LLM verdicts merged:              {verdict_match} rows")
    print(f"  Output: {OUT_XLSX}")


if __name__ == "__main__":
    main()
