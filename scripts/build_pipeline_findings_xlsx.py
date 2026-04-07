#!/usr/bin/env python3
"""
Build clean pipeline results from PDF-first preregistration detection outputs.

Inputs:
  - scan CSV
  - enriched links CSV
  - LLM verdict CSV

Outputs:
  - output/results.csv
  - output/results.xlsx
"""

import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from urllib.parse import urlparse

from path_utils import resolve_existing_path, resolve_output_path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not found - run: pip install openpyxl")


ROOT = Path(__file__).parent.parent
OUTPUT_DIR = ROOT / "output"
DEFAULT_SCAN = OUTPUT_DIR / "pdf_scan_results.csv"
FALLBACK_SCAN = OUTPUT_DIR / "pdf_scan_results_v2.csv"
DEFAULT_LINKS_CSV = OUTPUT_DIR / "pdf_scan_prereg_links_dedup.csv"
FALLBACK_LINKS_CSV = OUTPUT_DIR / "pdf_scan_prereg_links.csv"
DEFAULT_VERDICTS_CSV = OUTPUT_DIR / "llm_verdicts.csv"
FALLBACK_VERDICTS_CSV = OUTPUT_DIR / "llm_gemini_verdicts.csv"
DEFAULT_OUT_XLSX = OUTPUT_DIR / "results.xlsx"
DEFAULT_OUT_CSV = OUTPUT_DIR / "results.csv"

CONFIDENCE_MAP = {"high": 0.9, "medium": 0.6, "low": 0.25}
FINAL_ACCEPTED_LINK_QUALITIES = {"VERIFIED", "DOI_CONFIRMED", "AUTHOR_CONFIRMED", "AI_LINK_CONFIRMED"}
CAUTIOUS_LINK_DOMAINS = {"osf.io", "clinicaltrials.gov"}
TRUST_DIRECT_PDF_LINK_DOMAINS = {"aspredicted.org"}
GENERIC_LINK_PATTERNS = [
    re.compile(r"^https?://(?:www\.)?aspredicted\.org/?$", re.I),
    re.compile(r"^https?://(?:www\.)?aspredicted\.org/blind/?$", re.I),
    re.compile(r"^https?://(?:www\.)?egap\.org/?$", re.I),
    re.compile(r"^https?://(?:www\.)?egap\.org/registration/?$", re.I),
    re.compile(r"^https?://(?:www\.)?osf\.io/?$", re.I),
    re.compile(r"^https?://(?:www\.)?osf\.io/(download|preprints|registries|search|meetings|institutions)/?$", re.I),
    re.compile(r"^https?://(?:www\.)?socialscienceregistry\.org/?$", re.I),
    re.compile(r"^https?://(?:www\.)?socialscienceregistry\.org/trials/?$", re.I),
    re.compile(r"^https?://(?:www\.)?socialscienceregistry\.org/trials/0/?$", re.I),
]

META_FILL = PatternFill("solid", fgColor="E8E8E8")
SCAN_FILL = PatternFill("solid", fgColor="FFF2CC")
LINK_FILL = PatternFill("solid", fgColor="FCE5CD")
AI_FILL = PatternFill("solid", fgColor="DDEEFF")
FINAL_FILL = PatternFill("solid", fgColor="D9EAD3")
BOLD = Font(bold=True)

COLUMNS = [
    ("filename", META_FILL, 28),
    ("pdf_path", META_FILL, 54),
    ("journal", META_FILL, 30),
    ("title", META_FILL, 60),
    ("doi", META_FILL, 34),
    ("text_source", META_FILL, 18),
    ("no_data", SCAN_FILL, 9),
    ("auto_prereg", SCAN_FILL, 12),
    ("auto_use_aearct", SCAN_FILL, 14),
    ("auto_use_osf", SCAN_FILL, 12),
    ("auto_use_aspredicted", SCAN_FILL, 16),
    ("auto_use_other", SCAN_FILL, 13),
    ("auto_link_prereg", SCAN_FILL, 44),
    ("auto_type_lab", SCAN_FILL, 11),
    ("auto_type_field", SCAN_FILL, 12),
    ("auto_type_online", SCAN_FILL, 13),
    ("auto_type_survey", SCAN_FILL, 13),
    ("auto_type_obs", SCAN_FILL, 11),
    ("all_found_links", LINK_FILL, 60),
    ("best_link_quality", LINK_FILL, 18),
    ("best_link_title", LINK_FILL, 40),
    ("author_match", LINK_FILL, 24),
    ("final_use_aearct", FINAL_FILL, 14),
    ("final_use_osf", FINAL_FILL, 12),
    ("final_use_aspredicted", FINAL_FILL, 16),
    ("final_use_other", FINAL_FILL, 13),
    ("final_link_url", FINAL_FILL, 42),
    ("final_link_decision", FINAL_FILL, 14),
    ("final_link_source", FINAL_FILL, 18),
    ("ai_prereg", AI_FILL, 12),
    ("ai_confidence", AI_FILL, 14),
    ("ai_evidence", AI_FILL, 50),
    ("ai_registry_url", AI_FILL, 42),
    ("ai_reasoning", AI_FILL, 60),
    ("ai_evidence_location", AI_FILL, 28),
    ("experiment", FINAL_FILL, 11),
    ("final_prereg_decision", FINAL_FILL, 18),
    ("final_prereg_source", FINAL_FILL, 18),
]

_LOCATION_PATTERNS = [
    ("footnote", re.compile(r"\bfootnote", re.I)),
    ("acknowledgments", re.compile(r"\backnowledg", re.I)),
    ("abstract", re.compile(r"\babstract\b", re.I)),
    ("introduction", re.compile(r"\bintroduction\b", re.I)),
    ("data section", re.compile(r"\bdata\s+(?:section|appendix|subsection)\b", re.I)),
    ("methods section", re.compile(r"\bmethods?\s*(?:section|subsection)?\b|methodology\b", re.I)),
    ("appendix", re.compile(r"\bappendix\b", re.I)),
    ("contract/PAP", re.compile(r"\bcontract\b|\bpre.?analysis\s+plan\b", re.I)),
]


def load_csv(path: Path, key_col: str = "filename") -> dict:
    if not path.exists():
        return {}
    out = {}
    with open(path, encoding="utf-8", errors="ignore", newline="") as f:
        for row in csv.DictReader(f):
            key = (row.get(key_col) or "").strip()
            if key:
                out[key] = row
    return out


def int_or_none(val):
    try:
        if val in (None, ""):
            return None
        return int(float(val))
    except (TypeError, ValueError):
        return None


def to_bool_or_none(val):
    if val is None:
        return None
    text = str(val).strip().lower()
    if text == "true":
        return True
    if text == "false":
        return False
    return None


def clean_text_or_none(val):
    if val is None:
        return None
    text = str(val).strip()
    return text or None


def split_links(raw: str | None) -> list[str]:
    if not raw:
        return []
    out = []
    seen = set()
    for part in str(raw).split(";"):
        text = part.strip().rstrip(".,;:")
        if text and text not in seen:
            seen.add(text)
            out.append(text)
    return out


def is_generic_registry_link(url: str | None) -> bool:
    text = (url or "").strip().rstrip(".,;:")
    if not text:
        return True
    return any(pattern.match(text) for pattern in GENERIC_LINK_PATTERNS)


def specific_links(raw: str | None) -> list[str]:
    return [link for link in split_links(raw) if not is_generic_registry_link(link)]


def accepted_registry_links(links: list[str], ai_prereg_bool, quality: str | None = None) -> list[str]:
    accepted = []
    for link in links:
        domain = link_domain(link)
        if domain in TRUST_DIRECT_PDF_LINK_DOMAINS:
            accepted.append(link)
            continue
        if domain in CAUTIOUS_LINK_DOMAINS:
            if ai_prereg_bool is True:
                accepted.append(link)
            continue
        if quality == "AI_LINK_REJECTED":
            continue
        if ai_prereg_bool is False:
            continue
        accepted.append(link)
    return accepted


def unique_preserve(items: list[str]) -> list[str]:
    seen = set()
    out = []
    for item in items:
        if item and item not in seen:
            seen.add(item)
            out.append(item)
    return out


def canonicalize_link(url: str) -> str:
    text = (url or "").strip().rstrip(".,;:").lower()
    if not text:
        return ""
    aea_match = re.search(r"(?:socialscienceregistry\.org/trials/|aearctr[-:])0*(\d+)", text)
    if aea_match:
        return f"aearctr:{int(aea_match.group(1))}"
    asp_match = re.search(r"aspredicted\.org/(?:blind\.php\?x=)?([a-z0-9]+)", text)
    if asp_match:
        return f"aspredicted:{asp_match.group(1)}"
    osf_match = re.search(r"osf\.io/(?:preprints/osf/)?([a-z0-9]+)", text)
    if osf_match:
        return f"osf:{osf_match.group(1)}"
    text = re.sub(r"^https?://", "", text)
    text = re.sub(r"^www\.", "", text)
    return text.rstrip("/")


def link_domain(url: str | None) -> str:
    text = (url or "").strip().rstrip(".,;:")
    if not text or "://" not in text:
        return ""
    return urlparse(text).netloc.lower().removeprefix("www.")


def normalized_title(text: str | None) -> str:
    title = (text or "").lower()
    title = re.sub(r"[^\w\s]", " ", title)
    return re.sub(r"\s+", " ", title).strip()


def prefer_longer_matching_title(current: str | None, candidate: str | None) -> str | None:
    if not candidate:
        return current
    if not current:
        return candidate
    cur_norm = normalized_title(current)
    cand_norm = normalized_title(candidate)
    if not cur_norm or not cand_norm:
        return current
    if cur_norm == cand_norm:
        return candidate if len(candidate) > len(current) else current
    shorter, longer = sorted((cur_norm, cand_norm), key=len)
    boundary_extension = (
        longer.startswith(shorter + " ")
        or longer.endswith(" " + shorter)
        or shorter.startswith(longer + " ")
        or shorter.endswith(" " + longer)
    )
    if boundary_extension and (len(shorter) / len(longer)) >= 0.55:
        return candidate if len(candidate) > len(current) else current
    return current


def extract_evidence_location(evidence: str | None, reasoning: str | None) -> str | None:
    combined = f"{evidence or ''} {reasoning or ''}".strip()
    if not combined:
        return None
    found = []
    for label, pattern in _LOCATION_PATTERNS:
        if pattern.search(combined) and label not in found:
            found.append(label)
    return "; ".join(found) if found else None


def ai_supports_prereg_without_link(ai_prereg_bool, ai_evidence, ai_reasoning, ai_registry_url) -> bool:
    if ai_prereg_bool is not True:
        return False
    if clean_text_or_none(ai_registry_url):
        return True
    combined = f"{ai_evidence or ''} {ai_reasoning or ''}".lower()
    strong_patterns = [
        r"\bpre-?registered\b",
        r"\bpreregistered\b",
        r"\bpre-?registration\b",
        r"\bregistered report\b",
        r"\bregistered an analysis plan\b",
        r"\bpre-?analysis plan\b.{0,140}\b(?:before|prior to|a priori)\b",
        r"\bas we preregistered\b",
    ]
    return any(re.search(pattern, combined, re.I | re.DOTALL) for pattern in strong_patterns)


def maybe_fetch_registry_title(current_title, candidate_url, paper_title, paper_doi, best_quality, cache):
    if current_title or not candidate_url:
        return current_title
    if (best_quality or "").strip() not in FINAL_ACCEPTED_LINK_QUALITIES:
        return current_title
    if candidate_url not in cache:
        try:
            from find_prereg_links import validate_link_quality
            fetched = validate_link_quality(candidate_url, paper_title or "", paper_doi or "")
            cache[candidate_url] = clean_text_or_none(fetched.get("registry_page_title"))
        except Exception:
            cache[candidate_url] = None
    fetched_title = cache.get(candidate_url)
    if not fetched_title:
        return current_title
    return prefer_longer_matching_title(current_title or paper_title, fetched_title)


def pick_final_link(auto_link_raw, all_found_raw, ai_registry_url, best_quality, ai_prereg_bool):
    auto_links = specific_links(auto_link_raw)
    enriched_links = specific_links(all_found_raw)
    ai_links = specific_links(ai_registry_url) if ai_prereg_bool is True else []
    quality = (best_quality or "").strip()

    accepted_auto = accepted_registry_links(auto_links, ai_prereg_bool, quality)

    accepted_enriched = []
    if enriched_links and quality in FINAL_ACCEPTED_LINK_QUALITIES:
        accepted_enriched = accepted_registry_links(enriched_links, ai_prereg_bool, quality)

    final_links = unique_preserve(accepted_auto + accepted_enriched + ai_links)
    if accepted_auto:
        return accepted_auto[0], 1, "pdf_text", final_links
    if accepted_enriched:
        return accepted_enriched[0], 1, "enrichment_verified", final_links
    if ai_links:
        return ai_links[0], 1, "ai_registry", final_links
    return None, 0, "none", []


def derive_platform_flags(auto_use_aearct, auto_use_osf, auto_use_asp, auto_use_other, final_links):
    final_aearct = 1 if auto_use_aearct == 1 else 0
    final_osf = 1 if auto_use_osf == 1 else 0
    final_asp = 1 if auto_use_asp == 1 else 0
    final_other = 1 if auto_use_other == 1 else 0
    for link in final_links:
        canon = canonicalize_link(link)
        if canon.startswith("aearctr:"):
            final_aearct = 1
        elif canon.startswith("osf:"):
            final_osf = 1
        elif canon.startswith("aspredicted:"):
            final_asp = 1
        elif any(domain in canon for domain in ("clinicaltrials.gov", "egap.org", "ridie")):
            final_other = 1
    return final_aearct, final_osf, final_asp, final_other


def write_sheet_header(ws, columns):
    for col_idx, (col_name, fill, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = BOLD
        cell.fill = fill
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_summary_sheets(wb, rows: list[dict]):
    summary_ws = wb.create_sheet("summary")
    summary_ws.append(["metric", "value"])
    summary_ws["A1"].font = BOLD
    summary_ws["B1"].font = BOLD
    metrics = [
        ("total_papers", len(rows)),
        ("final_prereg_1", sum(1 for r in rows if r["final_prereg_decision"] == 1)),
        ("final_link_1", sum(1 for r in rows if r["final_link_decision"] == 1)),
        ("experiment_1", sum(1 for r in rows if r["experiment"] == 1)),
        ("ai_prereg_true", sum(1 for r in rows if r["ai_prereg"] == "True")),
        ("best_link_verified", sum(1 for r in rows if r["best_link_quality"] in FINAL_ACCEPTED_LINK_QUALITIES)),
    ]
    for metric, value in metrics:
        summary_ws.append([metric, value])

    journal_ws = wb.create_sheet("journal_summary")
    headers = ["journal", "total_papers", "experiment_papers", "final_prereg_papers", "final_link_papers"]
    journal_ws.append(headers)
    for cell in journal_ws[1]:
        cell.font = BOLD
    stats_by_journal = defaultdict(lambda: {"total": 0, "experiment": 0, "final_prereg": 0, "final_link": 0})
    for row in rows:
        journal = row["journal"] or "(missing)"
        stats = stats_by_journal[journal]
        stats["total"] += 1
        if row["experiment"] == 1:
            stats["experiment"] += 1
        if row["final_prereg_decision"] == 1:
            stats["final_prereg"] += 1
        if row["final_link_decision"] == 1:
            stats["final_link"] += 1
    for journal in sorted(stats_by_journal):
        stats = stats_by_journal[journal]
        journal_ws.append([journal, stats["total"], stats["experiment"], stats["final_prereg"], stats["final_link"]])


def main():
    parser = argparse.ArgumentParser(description="Build clean pipeline results files")
    parser.add_argument("--scan", type=str, default=None, help=f"Path to scan CSV (default: {DEFAULT_SCAN})")
    parser.add_argument("--links", type=str, default=None, help=f"Path to enriched links CSV (default: {DEFAULT_LINKS_CSV})")
    parser.add_argument("--verdicts", type=str, default=None, help=f"Path to LLM verdict CSV (default: {DEFAULT_VERDICTS_CSV})")
    parser.add_argument("--output-xlsx", type=str, default=None, help=f"Path to output workbook (default: {DEFAULT_OUT_XLSX})")
    parser.add_argument("--output-csv", type=str, default=None, help=f"Path to output CSV (default: {DEFAULT_OUT_CSV})")
    args = parser.parse_args()

    scan_csv = resolve_existing_path(args.scan, DEFAULT_SCAN, "scan CSV", fallbacks=[FALLBACK_SCAN])
    links_csv = resolve_existing_path(args.links, DEFAULT_LINKS_CSV, "enriched links CSV", fallbacks=[FALLBACK_LINKS_CSV], required=False)
    verdicts_csv = resolve_existing_path(
        args.verdicts,
        DEFAULT_VERDICTS_CSV,
        "LLM verdict CSV",
        fallbacks=[FALLBACK_VERDICTS_CSV],
        required=False,
    )
    out_xlsx = resolve_output_path(args.output_xlsx, DEFAULT_OUT_XLSX)
    out_csv = resolve_output_path(args.output_csv, DEFAULT_OUT_CSV)

    scan = load_csv(scan_csv)
    links = load_csv(links_csv)
    verdicts = load_csv(verdicts_csv)

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "pipeline_results"
    write_sheet_header(out_ws, COLUMNS)
    out_ws.freeze_panes = "A2"

    rows_for_summary = []
    title_cache = {}

    for filename, scan_row in sorted(scan.items()):
        link_row = links.get(filename, {})
        verdict_row = verdicts.get(filename, {})

        title = clean_text_or_none(link_row.get("title_guess"))
        best_link_title = clean_text_or_none(link_row.get("best_link_title"))
        title = prefer_longer_matching_title(title, best_link_title)
        doi = clean_text_or_none(link_row.get("doi_from_pdf"))
        pdf_path = clean_text_or_none(scan_row.get("pdf_path"))
        text_source = clean_text_or_none(scan_row.get("text_source"))

        no_data = int_or_none(scan_row.get("auto_no_data"))
        auto_prereg = int_or_none(scan_row.get("auto_prereg"))
        auto_use_aearct = int_or_none(scan_row.get("auto_use_aearct"))
        auto_use_osf = int_or_none(scan_row.get("auto_use_osf"))
        auto_use_asp = int_or_none(scan_row.get("auto_use_aspredicted"))
        auto_use_other = int_or_none(scan_row.get("auto_use_other"))
        auto_link_prereg = clean_text_or_none(scan_row.get("auto_link_prereg"))
        auto_type_lab = int_or_none(scan_row.get("auto_type_lab"))
        auto_type_field = int_or_none(scan_row.get("auto_type_field"))
        auto_type_online = int_or_none(scan_row.get("auto_type_online"))
        auto_type_survey = int_or_none(scan_row.get("auto_type_survey"))
        auto_type_obs = int_or_none(scan_row.get("auto_type_obs"))
        if auto_type_obs == 1 and any(v == 1 for v in (auto_type_lab, auto_type_field, auto_type_online)):
            auto_type_obs = 0

        all_found_links = clean_text_or_none(link_row.get("all_found_links"))
        best_link_quality = clean_text_or_none(link_row.get("best_link_quality"))
        author_match = clean_text_or_none(link_row.get("author_match"))
        if not auto_link_prereg:
            auto_link_prereg = clean_text_or_none(link_row.get("auto_link_prereg"))

        ai_prereg_bool = to_bool_or_none(verdict_row.get("llm_prereg"))
        ai_prereg = "True" if ai_prereg_bool is True else "False" if ai_prereg_bool is False else None
        conf_str = (clean_text_or_none(verdict_row.get("llm_confidence")) or "").lower()
        ai_confidence = CONFIDENCE_MAP.get(conf_str)
        ai_evidence = clean_text_or_none(verdict_row.get("llm_evidence"))
        ai_registry_url = clean_text_or_none(verdict_row.get("llm_registry_url"))
        ai_reasoning = clean_text_or_none(verdict_row.get("llm_reasoning"))
        ai_evidence_location = extract_evidence_location(ai_evidence, ai_reasoning)

        if ai_registry_url and not all_found_links:
            all_found_links = ai_registry_url
        elif ai_registry_url:
            all_found_links = "; ".join(unique_preserve(split_links(all_found_links) + split_links(ai_registry_url)))

        final_link_url, final_link_decision, final_link_source, final_links = pick_final_link(
            auto_link_prereg,
            all_found_links,
            ai_registry_url,
            best_link_quality,
            ai_prereg_bool,
        )
        best_link_title = maybe_fetch_registry_title(
            best_link_title,
            final_link_url,
            title,
            doi,
            best_link_quality,
            title_cache,
        )
        title = prefer_longer_matching_title(title, best_link_title)

        auto_use_aearct, auto_use_osf, auto_use_asp, auto_use_other = derive_platform_flags(
            auto_use_aearct, auto_use_osf, auto_use_asp, auto_use_other, final_links
        )
        final_use_aearct, final_use_osf, final_use_asp, final_use_other = derive_platform_flags(
            auto_use_aearct, auto_use_osf, auto_use_asp, auto_use_other, final_links
        )

        experiment = 1 if any(v == 1 for v in (auto_type_lab, auto_type_field, auto_type_online, auto_type_survey, auto_type_obs)) else 0
        ai_only_prereg = ai_supports_prereg_without_link(ai_prereg_bool, ai_evidence, ai_reasoning, ai_registry_url)
        if final_link_decision == 1 and ai_only_prereg:
            final_prereg_source = "link+ai"
        elif final_link_decision == 1:
            final_prereg_source = f"link:{final_link_source}"
        elif ai_only_prereg:
            final_prereg_source = "ai_only"
        else:
            final_prereg_source = "none"
        final_prereg_decision = 1 if (final_link_decision == 1 or ai_only_prereg) else 0

        row_dict = {
            "filename": filename,
            "pdf_path": pdf_path,
            "journal": scan_row.get("journal") or None,
            "title": title,
            "doi": doi,
            "text_source": text_source,
            "no_data": no_data,
            "auto_prereg": auto_prereg,
            "auto_use_aearct": auto_use_aearct,
            "auto_use_osf": auto_use_osf,
            "auto_use_aspredicted": auto_use_asp,
            "auto_use_other": auto_use_other,
            "auto_link_prereg": auto_link_prereg,
            "auto_type_lab": auto_type_lab,
            "auto_type_field": auto_type_field,
            "auto_type_online": auto_type_online,
            "auto_type_survey": auto_type_survey,
            "auto_type_obs": auto_type_obs,
            "all_found_links": all_found_links,
            "best_link_quality": best_link_quality,
            "best_link_title": best_link_title,
            "author_match": author_match,
            "final_use_aearct": final_use_aearct,
            "final_use_osf": final_use_osf,
            "final_use_aspredicted": final_use_asp,
            "final_use_other": final_use_other,
            "final_link_url": final_link_url,
            "final_link_decision": final_link_decision,
            "final_link_source": final_link_source,
            "ai_prereg": ai_prereg,
            "ai_confidence": ai_confidence,
            "ai_evidence": ai_evidence,
            "ai_registry_url": ai_registry_url,
            "ai_reasoning": ai_reasoning,
            "ai_evidence_location": ai_evidence_location,
            "experiment": experiment,
            "final_prereg_decision": final_prereg_decision,
            "final_prereg_source": final_prereg_source,
        }
        out_ws.append([row_dict[name] for name, _, _ in COLUMNS])
        rows_for_summary.append(row_dict)

    out_ws.auto_filter.ref = out_ws.dimensions
    write_summary_sheets(out_wb, rows_for_summary)
    out_wb.save(str(out_xlsx))

    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[name for name, _, _ in COLUMNS])
        writer.writeheader()
        writer.writerows(rows_for_summary)

    print(f"Done - {len(rows_for_summary)} papers written")
    print(f"  final_prereg_decision=1 : {sum(1 for r in rows_for_summary if r['final_prereg_decision'] == 1)}")
    print(f"  final_link_decision=1   : {sum(1 for r in rows_for_summary if r['final_link_decision'] == 1)}")
    print(f"  experiment=1            : {sum(1 for r in rows_for_summary if r['experiment'] == 1)}")
    print(f"  CSV  : {out_csv}")
    print(f"  XLSX : {out_xlsx}")


if __name__ == "__main__":
    main()
