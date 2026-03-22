#!/usr/bin/env python3
"""
Post-process the latest new_findings workbook into a cleaner review workbook.

This script uses ONLY the latest `output/new_findings_*.xlsx` file as input.
It does not read any CSV sources.
"""

import re
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not found - run: pip install openpyxl")


ROOT = Path(__file__).parent.parent
OUTPUT_DIR = ROOT / "output"

FINAL_ACCEPTED_LINK_QUALITIES = {
    "VERIFIED",
    "DOI_CONFIRMED",
    "AUTHOR_CONFIRMED",
    "AI_LINK_CONFIRMED",
}

SOURCE_FILL = PatternFill("solid", fgColor="E8E8E8")
FINAL_FILL = PatternFill("solid", fgColor="D9EAD3")
SUMMARY_FILL = PatternFill("solid", fgColor="DDEEFF")
BOLD = Font(bold=True)


def latest_new_findings() -> Path:
    candidates = sorted(OUTPUT_DIR.glob("new_findings_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError("No output/new_findings_*.xlsx file found")
    return candidates[0]


def split_links(raw) -> list[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def unique_preserve(items: list[str]) -> list[str]:
    seen = set()
    out = []
    for item in items:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def to_bool_or_none(val):
    if val is None:
        return None
    text = str(val).strip().lower()
    if text == "true":
        return True
    if text == "false":
        return False
    return None


def canonicalize_link(url: str) -> str:
    text = (url or "").strip().lower()
    if not text:
        return ""

    aea_match = re.search(r"(?:socialscienceregistry\.org/trials/|aearctr[-:])0*(\d+)", text)
    if aea_match:
        return f"aearctr:{int(aea_match.group(1))}"

    asp_match = re.search(r"aspredicted\.org/(?:blind\.php\?x=)?([a-z0-9]+)", text)
    if asp_match:
        return f"aspredicted:{asp_match.group(1)}"

    osf_match = re.search(r"osf\.io/(?:preprints/osf/)?([a-z0-9]+)", text)
    if osf_match:
        return f"osf:{osf_match.group(1)}"

    text = re.sub(r"^https?://", "", text)
    text = re.sub(r"^www\.", "", text)
    text = text.rstrip("/")
    return text


def derive_platform_flags(row: dict, final_links: list[str]) -> tuple[int, int, int, int]:
    final_aearct = 1 if row.get("auto_use_aearct") == 1 else 0
    final_osf = 1 if row.get("auto_use_osf") == 1 else 0
    final_asp = 1 if row.get("auto_use_aspredicted") == 1 else 0
    final_other = 1 if row.get("auto_use_other") == 1 else 0

    for link in final_links:
        canon = canonicalize_link(link)
        if canon.startswith("aearctr:"):
            final_aearct = 1
        elif canon.startswith("osf:"):
            final_osf = 1
        elif canon.startswith("aspredicted:"):
            final_asp = 1
        elif any(domain in canon for domain in ("clinicaltrials.gov", "egap.org", "ridie")):
            final_other = 1

    return final_aearct, final_osf, final_asp, final_other


def pick_final_link(row: dict) -> tuple[str | None, int, str, list[str]]:
    auto_links = split_links(row.get("auto_link_prereg"))
    enriched_links = split_links(row.get("all_found_links"))
    ai_links = split_links(row.get("ai_registry_url"))
    best_quality = str(row.get("best_link_quality") or "").strip()

    accepted_enriched = enriched_links if best_quality in FINAL_ACCEPTED_LINK_QUALITIES else []
    final_links = unique_preserve(auto_links + accepted_enriched + ai_links)

    if auto_links:
        return auto_links[0], 1, "pdf_text", final_links
    if accepted_enriched:
        return accepted_enriched[0], 1, "enrichment_verified", final_links
    if ai_links:
        return ai_links[0], 1, "ai_registry", final_links
    return None, 0, "none", []


def build_summary_sheet(wb, rows: list[dict]):
    ws = wb.create_sheet("summary")
    headers = ["metric", "value"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = BOLD
        cell.fill = SUMMARY_FILL

    metrics = [
        ("total_rows", len(rows)),
        ("auto_prereg_1", sum(1 for r in rows if r["auto_prereg"] == 1)),
        ("auto_link_prereg_nonempty", sum(1 for r in rows if r["auto_link_prereg"])),
        ("all_found_links_nonempty", sum(1 for r in rows if r["all_found_links"])),
        ("final_link_decision_1", sum(1 for r in rows if r["final_link_decision"] == 1)),
        ("final_prereg_decision_1", sum(1 for r in rows if r["final_prereg_decision"] == 1)),
        ("experiment_1", sum(1 for r in rows if r["experiment"] == 1)),
    ]

    for metric, value in metrics:
        ws.append([metric, value])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14


def build_journal_summary_sheet(wb, rows: list[dict]):
    ws = wb.create_sheet("journal_summary")
    headers = [
        "journal",
        "total_papers",
        "experiment_papers",
        "final_prereg_papers",
        "final_prereg_and_experiment",
        "prereg_share_all",
        "prereg_share_experiment",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = BOLD
        cell.fill = SUMMARY_FILL

    stats = {}
    for row in rows:
        journal = row.get("journal") or "(missing)"
        if journal not in stats:
            stats[journal] = {
                "total": 0,
                "experiment": 0,
                "final_prereg": 0,
                "exp_prereg": 0,
            }
        stats[journal]["total"] += 1
        if row["experiment"] == 1:
            stats[journal]["experiment"] += 1
        if row["final_prereg_decision"] == 1:
            stats[journal]["final_prereg"] += 1
        if row["experiment"] == 1 and row["final_prereg_decision"] == 1:
            stats[journal]["exp_prereg"] += 1

    for journal in sorted(stats):
        s = stats[journal]
        share_all = s["final_prereg"] / s["total"] if s["total"] else None
        share_exp = s["exp_prereg"] / s["experiment"] if s["experiment"] else None
        ws.append([
            journal,
            s["total"],
            s["experiment"],
            s["final_prereg"],
            s["exp_prereg"],
            share_all,
            share_exp,
        ])

    widths = (35, 12, 16, 18, 24, 16, 22)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def main():
    source_path = latest_new_findings()
    out_path = OUTPUT_DIR / f"new_findings_final_{date.today()}.xlsx"

    print(f"Using source workbook: {source_path.name}")

    wb_in = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    ws_in = wb_in.active
    source_headers = [c.value for c in next(ws_in.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(source_headers)}

    final_columns = [
        "final_use_aearct",
        "final_use_osf",
        "final_use_aspredicted",
        "final_use_other",
        "final_link_url",
        "final_link_decision",
        "final_link_source",
        "experiment",
        "final_prereg_decision",
        "final_prereg_source",
    ]

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "pipeline_findings_final"

    all_headers = source_headers + final_columns
    for col_idx, name in enumerate(all_headers, start=1):
        cell = ws_out.cell(row=1, column=col_idx, value=name)
        cell.font = BOLD
        cell.fill = FINAL_FILL if name in final_columns else SOURCE_FILL
        ws_out.column_dimensions[get_column_letter(col_idx)].width = 18

    rows_for_summary = []

    for row in ws_in.iter_rows(min_row=2, values_only=True):
        row_dict = {name: row[idx[name]] for name in source_headers}

        if row_dict.get("auto_type_obs") == 1 and any(
            row_dict.get(key) == 1 for key in ("auto_type_lab", "auto_type_field", "auto_type_online")
        ):
            row_dict["auto_type_obs"] = 0

        final_link_url, final_link_decision, final_link_source, final_links = pick_final_link(row_dict)
        final_use_aearct, final_use_osf, final_use_asp, final_use_other = derive_platform_flags(row_dict, final_links)

        experiment = 1 if any(
            row_dict.get(key) == 1
            for key in ("auto_type_lab", "auto_type_field", "auto_type_online", "auto_type_survey", "auto_type_obs")
        ) else 0

        ai_prereg_bool = to_bool_or_none(row_dict.get("ai_prereg"))
        final_prereg_decision = 1 if (final_link_decision == 1 or ai_prereg_bool is True) else 0

        if final_link_decision == 1 and ai_prereg_bool is True:
            final_prereg_source = "link+ai"
        elif final_link_decision == 1:
            final_prereg_source = f"link:{final_link_source}"
        elif ai_prereg_bool is True:
            final_prereg_source = "ai_only"
        else:
            final_prereg_source = "none"

        row_dict["final_use_aearct"] = final_use_aearct
        row_dict["final_use_osf"] = final_use_osf
        row_dict["final_use_aspredicted"] = final_use_asp
        row_dict["final_use_other"] = final_use_other
        row_dict["final_link_url"] = final_link_url
        row_dict["final_link_decision"] = final_link_decision
        row_dict["final_link_source"] = final_link_source
        row_dict["experiment"] = experiment
        row_dict["final_prereg_decision"] = final_prereg_decision
        row_dict["final_prereg_source"] = final_prereg_source

        ws_out.append([row_dict.get(name) for name in all_headers])
        rows_for_summary.append(row_dict)

    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = ws_out.dimensions

    build_summary_sheet(wb_out, rows_for_summary)
    build_journal_summary_sheet(wb_out, rows_for_summary)
    wb_out.save(out_path)
    wb_in.close()

    print(f"Output: {out_path}")
    print(f"  final_link_decision=1   : {sum(1 for r in rows_for_summary if r['final_link_decision'] == 1)}")
    print(f"  final_prereg_decision=1 : {sum(1 for r in rows_for_summary if r['final_prereg_decision'] == 1)}")
    print(f"  experiment=1            : {sum(1 for r in rows_for_summary if r['experiment'] == 1)}")


if __name__ == "__main__":
    main()
