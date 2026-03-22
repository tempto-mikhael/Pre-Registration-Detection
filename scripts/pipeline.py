"""
pipeline.py
-----------
Main pipeline for the ERC automation project.

For each paper in the xlsx spreadsheet:
  1.  Reconstruct or look up the DOI.
  2.  Fetch metadata (title, abstract, OA PDF URL) from OpenAlex / CrossRef.
  3.  If an OA PDF URL is available, download and extract full text.
  4.  Run automated checks against the instructions:
        • no_data  — does the paper use empirical data?
        • prereg   — does the paper mention pre-registration?
        • platform — which registry (AEA-RCT, OSF, AsPredicted, other)?
        • link_prereg — any pre-registration URL in the text?
        • type_*   — lab / field / online / survey / obs / replication?
  5.  Write results to  output/results.csv  (resumable).

Usage:
  python scripts/pipeline.py [--start ROW] [--end ROW] [--sample N]

  --start / --end  : 1-based row numbers in the xlsx (data starts at row 3)
  --sample N       : process at most N rows (useful for testing)
  --email EMAIL    : your e-mail for polite-pool API access
"""

import argparse
import csv
import io
import os
import re
import sys
import time
from pathlib import Path

# Allow running from the project root
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm

from doi_resolver import resolve_doi
from api_client import fetch_metadata, CONTACT_EMAIL

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
XLSX_PATH    = PROJECT_ROOT / "journal_articles_with_pap_2025-03-14.xlsx"
OUTPUT_DIR   = PROJECT_ROOT / "output"
OUTPUT_CSV   = OUTPUT_DIR / "results.csv"
PDF_DIR      = OUTPUT_DIR / "oa_pdfs"

OUTPUT_DIR.mkdir(exist_ok=True)
PDF_DIR.mkdir(exist_ok=True)

# ── Keywords (mirror the spreadsheet's key_* columns) ────────────────────────
# ---------------------------------------------------------------------------
# Keyword definitions
# Each entry is either a plain substring (for unambiguous multi-word phrases)
# or a compiled regex (for short/ambiguous tokens that need word boundaries).
# ---------------------------------------------------------------------------

# --- Pre-registration keyword patterns ---
# Multi-word phrases: safe as plain substring (very unlikely to appear by accident)
PREREG_PHRASES = [
    "analysis plan",
    "pre-analysis plan",
    "pre-analysis-plan",
    "pre analysis plan",
    "preanalysis plan",
    "pre-registration",
    "preregistration",
    "pre registration",
    "pre-register",
    "preregister",
    "pre-registered",
    "preregistered",
    "pre-registering",
    "aea rct",
    "aearctr-",                    # AEA RCT ID prefix (e.g. AEARCTR-0002106)
    "socialscienceregistry.org",
    "open science framework",
    "aspredicted.org",
    "osf.io",
    "clinicaltrials.gov",
    "egap.org",
    "ridie",
]

# Short tokens that need \\b word-boundary protection
PREREG_WORD_TOKENS = [
    r"\bpap\b",           # Pre-Analysis Plan abbreviation — word boundary prevents "paper"
    r"\baearct\b",        # AEA RCT registry abbreviation
    r"\bosf\b",           # Open Science Framework — word boundary prevents "professor"
    r"\begap\b",
    r"\baspredicted\b",
]

COMPILED_PREREG_WORDS = [re.compile(p, re.IGNORECASE) for p in PREREG_WORD_TOKENS]

# --- Experiment keywords (plain substrings, long enough to be unambiguous) ---
EXPERIMENT_PHRASES = [
    "field experiment",
    "laboratory experiment",
    "lab experiment",
    "online experiment",
    "randomized experiment",
    "randomized controlled trial",
    "randomized control trial",
    "randomised controlled trial",
    "randomized evaluation",
    "randomized trial",
    "rct ",           # trailing space avoids partial matches like "aircraft"
    " rct",           # leading space
    "(rct)",
    "intervention",
]

EXPERIMENT_WORD_TOKENS = [
    r"\bexperiment\b",
    r"\blaboratory\b",
]

COMPILED_EXPERIMENT_WORDS = [re.compile(p, re.IGNORECASE) for p in EXPERIMENT_WORD_TOKENS]

# --- Data presence keywords ---
DATA_PHRASES = [
    "regression", "coefficient", "observational data",
    "administrative data", "panel data", "cross-section",
    "survey data", "empirical", "estimation",
]

DATA_WORD_TOKENS = [r"\bdata\b", r"\bsample\b", r"\bobservations\b", r"\bsurvey\b"]
COMPILED_DATA_WORDS = [re.compile(p, re.IGNORECASE) for p in DATA_WORD_TOKENS]

# Pre-registration URL patterns
PREREG_URL_PATTERNS = [
    r"https?://(?:www\.)?socialscienceregistry\.org/trials/\d+",
    r"AEARCTR-\d+",
    r"https?://(?:www\.)?osf\.io/[A-Za-z0-9]+",
    r"https?://(?:www\.)?aspredicted\.org/\S+",
    r"https?://(?:www\.)?clinicaltrials\.gov/\S+",
    r"https?://(?:www\.)?egap\.org/\S+",
    r"\bAsPredicted\s*#\s*\d+",
    r"\bAsPredicted\s*\([^)]*\d[^)]*\)",
]

COMPILED_PREREG_URLS = [re.compile(p, re.IGNORECASE) for p in PREREG_URL_PATTERNS]

# Voter / legal contexts that trigger pre-reg keywords but are NOT research pre-reg
PREREG_VOTER_PHRASES = [
    "preregistration law",
    "pre-registration law",
    "preregistration statute",
    "voter preregistration",
    "voting preregistration",
    "youth preregistration",
    "election preregistration",
    "preregistration requirement",
    "preregistration program",
    "preregistration policy",
]

# ── Output columns ─────────────────────────────────────────────────────────────
CSV_FIELDS = [
    "row_num", "id", "journal", "pdf_filename", "file_name",
    "title_xlsx",
    # resolved
    "doi", "resolved_via",
    # fetched
    "title_fetched", "pub_year", "abstract",
    "oa_pdf_url", "oa_pdf_downloaded", "text_source",
    # automated checks
    "auto_no_data",       # 1 = likely no empirical data
    "auto_prereg",        # 1 = pre-registration likely present
    "auto_use_aearct",
    "auto_use_osf",
    "auto_use_aspredicted",
    "auto_use_other",
    "auto_link_prereg",   # extracted URL(s), semicolon-separated
    "auto_type_lab",
    "auto_type_field",
    "auto_type_online",
    "auto_type_survey",
    "auto_type_obs",
    # existing coded values from xlsx (for comparison)
    "xlsx_no_data", "xlsx_prereg",
    "xlsx_use_aearct", "xlsx_use_osf", "xlsx_use_aspredicted", "xlsx_use_other",
    "xlsx_link_prereg",
]

# ── Text extraction from downloaded PDF ──────────────────────────────────────

BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/pdf,*/*",
}


def is_pdf(content: bytes) -> bool:
    """Check if bytes look like a PDF (marker may not be at byte 0 on some servers)."""
    return b"%PDF" in content[:1024]


def download_pdf(url: str, dest_path: Path) -> bytes | None:
    """Download a PDF from url; cache locally at dest_path."""
    if dest_path.exists():
        data = dest_path.read_bytes()
        return data if is_pdf(data) else None
    try:
        r = requests.get(url, timeout=30, headers=BROWSER_HEADERS,
                         allow_redirects=True)
        if r.status_code == 200 and is_pdf(r.content):
            dest_path.write_bytes(r.content)
            return r.content
    except requests.RequestException:
        pass
    return None


def try_download_any(candidates: list, pdf_filename: str) -> tuple[bytes | None, str | None]:
    """
    Try each candidate URL in order until a real PDF is downloaded.
    Returns (pdf_bytes, url_that_worked) or (None, None).
    """
    for url in candidates:
        # Skip URLs that are unlikely to be direct PDFs
        # (e.g. bare doi.org landing pages)
        if url.startswith("https://doi.org/") and "/" not in url[16:]:
            continue
        safe_name = re.sub(r"[^\w]", "_", pdf_filename or str(id(url))) + ".pdf"
        dest = PDF_DIR / safe_name
        data = download_pdf(url, dest)
        if data:
            return data, url
    return None, None

def extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str | None:
    """Extract text from PDF bytes using PyMuPDF, falling back to pdfminer."""
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        text = "\n".join(page.get_text() for page in doc)
        doc.close()
        return text if text.strip() else None
    except ImportError:
        pass
    try:
        from pdfminer.high_level import extract_text as pm_extract
        text = pm_extract(io.BytesIO(pdf_bytes))
        return text if text.strip() else None
    except ImportError:
        pass
    return None
# Registry domains to search for in anchor hrefs
_REGISTRY_DOMAINS = [
    "socialscienceregistry.org", "osf.io", "aspredicted.org",
    "clinicaltrials.gov", "egap.org", "ridie.3ieimpact.org",
]


def _unique(lst: list) -> list:
    seen = set(); return [x for x in lst if not (x in seen or seen.add(x))]


def scrape_landing_page_links(doi: str) -> list[str]:
    """
    Fetch the journal landing page and extract pre-registration registry URLs.
    Called only when auto_prereg=1 and no link was found in the text.
    """
    if not doi or doi.startswith("TITLE_SLUG:") or doi.startswith("PII:"):
        return []

    url = (f"https://www.aeaweb.org/articles?id={doi}"
           if doi.startswith("10.1257/") else f"https://doi.org/{doi}")
    try:
        resp = requests.get(url, timeout=20, headers=BROWSER_HEADERS, allow_redirects=True)
        if resp.status_code != 200:
            return []
        html = resp.text
        soup = BeautifulSoup(html, "lxml")
        links = []
        # 1. All <a href> pointing to registry domains
        for tag in soup.find_all("a", href=True):
            href = tag["href"].strip()
            if any(d in href for d in _REGISTRY_DOMAINS):
                links.append(href)
        # 2. Regex scan of full HTML for IDs / bare URLs
        links.extend(extract_prereg_urls(html))
        # 3. Structured metadata tags
        for meta in soup.find_all("meta"):
            content = meta.get("content", "")
            if any(d in content for d in _REGISTRY_DOMAINS):
                links.extend(extract_prereg_urls(content))
        return _unique(links)
    except Exception:
        return []


def phrase_hit(text: str, phrases: list[str]) -> bool:
    """Return True if any plain phrase appears in text (case-insensitive)."""
    tl = text.lower()
    return any(p.lower() in tl for p in phrases)


def regex_hit(text: str, patterns: list) -> bool:
    """Return True if any compiled regex matches text."""
    return any(p.search(text) for p in patterns)


def extract_prereg_urls(text: str) -> list[str]:
    """Extract all pre-registration URLs / IDs from text.

    AEARCTR-NNNNN IDs are converted to full socialscienceregistry.org URLs.
    """
    found = []
    for pat in COMPILED_PREREG_URLS:
        for match in pat.findall(text):
            # Convert bare AEARCTR-NNNNN to a full URL
            if re.match(r"^AEARCTR-\d+$", match, re.IGNORECASE):
                trial_num = int(re.search(r"\d+", match).group())
                found.append(f"https://www.socialscienceregistry.org/trials/{trial_num}")
            else:
                found.append(match)
    seen = set()
    return [x for x in found if not (x in seen or seen.add(x))]


def auto_check(text: str) -> dict:
    """
    Run automated checks on extracted text.
    Uses phrase matching for unambiguous multi-word terms and
    word-boundary regex for short/ambiguous tokens.
    """
    # -- Has empirical data? --------------------------------------------------
    has_data = phrase_hit(text, DATA_PHRASES) or regex_hit(text, COMPILED_DATA_WORDS)
    has_exp  = phrase_hit(text, EXPERIMENT_PHRASES) or regex_hit(text, COMPILED_EXPERIMENT_WORDS)
    no_data  = 0 if (has_data or has_exp) else 1

    # -- Pre-registration? ----------------------------------------------------
    has_prereg_phrase = phrase_hit(text, PREREG_PHRASES)
    has_prereg_token  = regex_hit(text, COMPILED_PREREG_WORDS)
    prereg = 1 if (has_prereg_phrase or has_prereg_token) else 0

    # Suppress false positives caused by voter/legal preregistration context:
    # If the ONLY reason we fired is generic "preregistration" / "pre-registration"
    # appearing in a voter/election context, cancel the hit.
    if prereg == 1 and phrase_hit(text, PREREG_VOTER_PHRASES):
        # Voter context is present; only keep prereg=1 if there is also a
        # registry-specific signal (URL, AEARCTR id, PAP, OSF, AsPredicted, etc.)
        registry_signal = (
            phrase_hit(text, ["analysis plan", "pre-analysis plan", "preanalysis plan",
                               "aearctr-", "socialscienceregistry.org",
                               "osf.io", "aspredicted.org", "clinicaltrials.gov",
                               "egap.org", "open science framework"])
            or regex_hit(text, [re.compile(r"AEARCTR-\d+", re.I),
                                 re.compile(r"\bpap\b", re.I),
                                 re.compile(r"\bosf\b", re.I)])
        )
        prereg = 1 if registry_signal else 0

    # -- Platform-specific ----------------------------------------------------
    use_aearct = 1 if (phrase_hit(text, ["aea rct", "aearctr-", "socialscienceregistry.org"])
                       or regex_hit(text, [re.compile(r"\baearct\b", re.I)])) else 0
    use_osf    = 1 if phrase_hit(text, ["open science framework", "osf.io"]) else 0
    use_asp    = 1 if phrase_hit(text, ["aspredicted.org", "aspredicted"]) else 0
    use_other  = 1 if phrase_hit(text, ["clinicaltrials.gov", "egap.org", "ridie"]) else 0

    prereg_urls = extract_prereg_urls(text)

    # -- Experiment type ------------------------------------------------------
    type_lab    = 1 if phrase_hit(text, ["laboratory experiment", "lab experiment",
                                         "laboratory setting", "lab setting"]) \
                       or regex_hit(text, [re.compile(r"\blaboratory\b", re.I)]) else 0
    type_field  = 1 if phrase_hit(text, ["field experiment", "randomized controlled trial",
                                         "randomized control trial", "rct ", " rct",
                                         "(rct)", "randomized evaluation"]) else 0
    type_online = 1 if phrase_hit(text, ["online experiment", "mechanical turk",
                                         "mturk", "prolific", "amazon turk"]) else 0
    type_survey = 1 if (regex_hit(text, [re.compile(r"\bsurvey\b", re.I)])
                        and not type_lab and not type_field and not type_online) else 0
    type_obs    = 1 if (phrase_hit(text, ["observational data", "administrative data",
                                          "panel data", "census data",
                                          "administrative records"])
                        and not type_lab and not type_field and not type_online) else 0

    return {
        "auto_no_data":          no_data,
        "auto_prereg":           prereg,
        "auto_use_aearct":       use_aearct,
        "auto_use_osf":          use_osf,
        "auto_use_aspredicted":  use_asp,
        "auto_use_other":        use_other,
        "auto_link_prereg":      "; ".join(prereg_urls),
        "auto_type_lab":         type_lab,
        "auto_type_field":       type_field,
        "auto_type_online":      type_online,
        "auto_type_survey":      type_survey,
        "auto_type_obs":         type_obs,
    }

# ── Load already-processed rows ───────────────────────────────────────────────

def load_done_rows(csv_path: Path) -> set[int]:
    """Return set of row_num already present in the output CSV."""
    done = set()
    if csv_path.exists():
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    done.add(int(row["row_num"]))
                except (KeyError, ValueError):
                    pass
    return done

# ── Main ──────────────────────────────────────────────────────────────────────

def load_xlsx(path: Path):
    """Load the workbook and return (worksheet, column_index_map)."""
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    # Build column letter → column name map from row 2 (actual headers)
    col_map = {}
    for cell in ws[2]:
        if cell.value:
            col_map[cell.column_letter] = cell.value
    # Inverse: name → column index (1-based)
    name_to_col = {cell.value: cell.column for cell in ws[2] if cell.value}
    return ws, name_to_col


def process_row(ws_row, name_to_col: dict, row_num: int) -> dict:
    """Process one row and return a result dict."""

    def cell(col_name):
        idx = name_to_col.get(col_name)
        return ws_row[idx - 1].value if idx else None

    record = {
        "row_num":      row_num,
        "id":           cell("id"),
        "journal":      cell("journal"),
        "pdf_filename": cell("pdf"),
        "file_name":    cell("file_name"),
        "title_xlsx":   cell("file_title"),
        # existing coded values
        "xlsx_no_data":         cell("no_data"),
        "xlsx_prereg":          cell("prereg"),
        "xlsx_use_aearct":      cell("use_aearct"),
        "xlsx_use_osf":         cell("use_osf"),
        "xlsx_use_aspredicted": cell("use_aspredicted"),
        "xlsx_use_other":       cell("use_other"),
        "xlsx_link_prereg":     cell("link_prereg"),
    }

    journal  = record["journal"] or ""
    pdf_file = record["pdf_filename"] or ""
    title    = record["title_xlsx"] or ""

    # Skip completely blank xlsx rows (trailing empty rows beyond actual data)
    if not journal and not pdf_file and not title:
        return None

    # ── Step 1: Resolve DOI ──────────────────────────────────────────────────
    raw_doi = resolve_doi(journal, pdf_file)

    doi        = None
    pii        = None
    title_slug = None

    if raw_doi:
        if raw_doi.startswith("PII:"):
            pii = raw_doi[4:]
        elif raw_doi.startswith("TITLE_SLUG:"):
            title_slug = raw_doi[11:]
        else:
            doi = raw_doi

    # ── Step 2: Fetch metadata ───────────────────────────────────────────────
    meta = fetch_metadata(
        doi=doi,
        pii=pii,
        title_slug=title_slug,
        title=title or None,
        journal=journal,
    )

    record["doi"]           = meta.get("doi") or doi
    record["resolved_via"]  = meta.get("source")
    record["title_fetched"] = meta.get("title")
    record["pub_year"]      = meta.get("pub_year")
    record["abstract"]      = (meta.get("abstract") or "")[:2000]  # truncate for CSV
    record["oa_pdf_url"]    = meta.get("oa_pdf_url")

    # ── Step 3: Attempt OA PDF download + full text extraction ───────────────
    full_text = meta.get("abstract") or ""        # seed with abstract
    oa_downloaded = 0
    oa_url_used = ""

    candidates = meta.get("oa_pdf_candidates") or []
    pdf_bytes, url_used = try_download_any(candidates, pdf_file or str(row_num))
    if pdf_bytes:
        extracted = extract_text_from_pdf_bytes(pdf_bytes)
        if extracted and len(extracted) > len(full_text):
            full_text = extracted
            oa_downloaded = 1
            oa_url_used = url_used or ""

    record["oa_pdf_downloaded"] = oa_downloaded
    record["oa_pdf_url"]        = oa_url_used or meta.get("oa_pdf_url") or ""
    record["text_source"]       = "full_pdf" if oa_downloaded else ("abstract" if full_text.strip() else "none")

    # ── Step 4: Automated checks (from text only — no xlsx keyword data used) ─
    if full_text.strip():
        checks = auto_check(full_text)
    else:
        checks = {k: "" for k in [
            "auto_no_data", "auto_prereg",
            "auto_use_aearct", "auto_use_osf", "auto_use_aspredicted", "auto_use_other",
            "auto_link_prereg",
            "auto_type_lab", "auto_type_field", "auto_type_online",
            "auto_type_survey", "auto_type_obs",
        ]}

    record.update(checks)

    # ── Step 5: If auto_prereg=1 and no link found yet, scrape landing page ──
    if checks.get("auto_prereg") == 1 and not checks.get("auto_link_prereg"):
        doi = record.get("doi", "") or ""
        lp_links = scrape_landing_page_links(doi)
        if lp_links:
            record["auto_link_prereg"] = "; ".join(lp_links)

    return record


def run(start_row: int = 3, end_row: int = None,
        sample: int = None, delay: float = 0.3,
        skip_prereg: bool = False):
    """Main entry point."""
    ws, name_to_col = load_xlsx(XLSX_PATH)

    max_data_row = ws.max_row
    end_row = min(end_row or max_data_row, max_data_row)

    done_rows = load_done_rows(OUTPUT_CSV)
    write_header = not OUTPUT_CSV.exists() or len(done_rows) == 0

    # Column index for xlsx_prereg (column J = index 10, 1-based)
    prereg_col_idx = name_to_col.get("prereg")  # 1-based

    rows_to_process = []
    skipped_prereg = 0
    for r in range(start_row, end_row + 1):
        if r in done_rows:
            continue
        if skip_prereg and prereg_col_idx:
            val = ws.cell(r, prereg_col_idx).value
            if val == 1:
                skipped_prereg += 1
                continue
        rows_to_process.append(r)

    if sample:
        rows_to_process = rows_to_process[:sample]

    print(f"Processing {len(rows_to_process)} rows "
          f"(rows {start_row}\u2013{end_row}, skipping {len(done_rows)} already done"
          + (f", skipping {skipped_prereg} already coded prereg=1" if skip_prereg else "")
          + ")")

    with open(OUTPUT_CSV, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
        if write_header:
            writer.writeheader()

        for row_num in tqdm(rows_to_process, unit="paper"):
            ws_row = ws[row_num]
            try:
                result = process_row(ws_row, name_to_col, row_num)
            except Exception as exc:
                print(f"\n[row {row_num}] ERROR: {exc}")
                result = {"row_num": row_num, "resolved_via": f"ERROR: {exc}"}

            if result is None:
                continue  # blank xlsx row — skip entirely

            writer.writerow(result)
            f.flush()
            time.sleep(delay)   # be polite to APIs

    print(f"\nDone. Results saved to: {OUTPUT_CSV}")


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="ERC Automation Pipeline")
    parser.add_argument("--start",  type=int, default=3,
                        help="First xlsx row to process (default: 3, i.e. first data row)")
    parser.add_argument("--end",    type=int, default=None,
                        help="Last xlsx row to process (default: last row)")
    parser.add_argument("--sample", type=int, default=None,
                        help="Process at most N rows (for testing)")
    parser.add_argument("--delay",  type=float, default=0.3,
                        help="Seconds to wait between API calls (default: 0.3)")
    parser.add_argument("--email",  type=str, default=None,
                        help="Your e-mail for polite-pool API access")
    parser.add_argument("--skip-prereg", action="store_true", default=False,
                        help="Skip rows already manually coded as prereg=1 in the xlsx")
    args = parser.parse_args()

    if args.email:
        import api_client
        api_client.CONTACT_EMAIL = args.email
        api_client.SESSION.headers.update(
            {"User-Agent": f"ercautomation/1.0 (mailto:{args.email})"}
        )

    run(start_row=args.start,
        end_row=args.end,
        sample=args.sample,
        delay=args.delay,
        skip_prereg=args.skip_prereg)
